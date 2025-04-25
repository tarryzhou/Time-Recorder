

'''------------------------------------V9---------------------------------------------------------'''

import tkinter as tk
from tkinter import messagebox
import time
from tkinter import ttk
from openpyxl import Workbook
import pandas as pd
from tkinter import StringVar
import os
from datetime import datetime
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from io import BytesIO

class TimerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("计时器")
        self.canvas = None          # 画布组件引用
        self.current_chart = "pie"  # 当前图表类型（pie/bar）
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 替换为系统支持的中文字体
        plt.rcParams['axes.unicode_minus'] = False     # 正确显示负号
        # 初始化用于保存图表图像的变量
        self.pie_img = None
        self.bar_img = None

        # 创建 Notebook 组件用于管理页签
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        '''--------------------------------------------------------------------计时器页签-----------------------------------------------------------------------'''
        # 创建计时器页签
        self.timer_tab = tk.Frame(self.notebook)
        self.notebook.add(self.timer_tab, text="计时器")

        # 初始化时间变量 （属性）
        self.start_time = None         #初始化开始按钮的状态，后续用于记录点击“开始计时”的起始时间。
        self.pause_start_time = None   #初始化暂停按钮的状态，后续用于记录点击“暂停计时”的起始时间
        self.pause_total_time = 0      #初始化从点击“暂停计时”到“恢复计时”累计的时间
        self.is_paused = False         #初始化暂停按钮的状态，后续用于标记和判断“暂停按钮”目前所在状态
        self.record_index = 1          #初始化行序号

        # 创建上半部分框架并分为左右区域
        top_frame = tk.Frame(self.timer_tab)
        top_frame.pack(pady=10)

        # 左边按钮区域
        left_frame = tk.Frame(top_frame)
        left_frame.pack(side=tk.LEFT, padx=10) #pady=10 表示在 top_frame 的上下两边各留出 10 个像素的空白空间

        # 创建按钮
        self.start_button = tk.Button(left_frame, text="开始计时", command=self.start_timer)
        self.start_button.pack(side=tk.LEFT, padx=5)

        self.pause_button = tk.Button(left_frame, text="暂停计时", command=self.pause_timer, state=tk.DISABLED,
                                        bg="red", fg="white")
        self.pause_button.pack(side=tk.LEFT, padx=5)

        self.end_button = tk.Button(left_frame, text="结束计时", command=self.end_timer, state=tk.DISABLED) #state=tk.DISABLED把按钮设置成了禁用状态，通常会呈现出灰色
        self.end_button.pack(side=tk.LEFT, padx=5)

        self.export_button = tk.Button(left_frame, text="导出Excel", command=self.export_to_excel, state=tk.DISABLED)
        self.export_button.pack(side=tk.LEFT, padx=5)

        # 右边备注下拉选择框区域
        right_frame = tk.Frame(top_frame)
        right_frame.pack(side=tk.RIGHT, padx=10)

        self.note_label = tk.Label(right_frame, text="备注")
        self.note_label.pack()
        self.todo_var = StringVar()
        self.note_entry = ttk.Combobox(right_frame, textvariable=self.todo_var, width=30)
        self.note_entry.pack()
        self.note_entry.bind("<Button-1>", self.update_combobox_on_click)

        # 创建表格和滚动条
        columns = ("序号", "开始时间", "结束时间", "暂停时长", "用工时长", "备注")
        self.timer_tree = ttk.Treeview(self.timer_tab, columns=columns, show="headings", height=10) # 列的数量和名称由columns指定，"headings" 表示只显示列标题（表头），而不显示默认的第一列
        for col in columns:                   #使用 for 循环遍历 columns 序列中的每个元素，针对表格的每一列执行两个操作：设置表头文本和设置列宽。
            self.timer_tree.heading(col, text=col)  #heading() 是 Treeview 组件的一个方法，用于设置指定列的表头信息
            self.timer_tree.column(col, width=120)  #用于设置指定列的属性

        self.scrollbar = ttk.Scrollbar(self.timer_tab, orient="vertical", command=self.timer_tree.yview) #创建一个垂直滚动条组件，关联到tree的垂直滚动操作
        self.timer_tree.configure(yscrollcommand=self.scrollbar.set)                                     #将tree的垂直滚动操作关联到 scrollbar
        self.timer_tree.bind("<MouseWheel>", self.on_mousewheel)

        self.timer_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=10) #tk.BOTH 表示组件将在水平和垂直方向上都填充可用空间，也就是 self.timer_tree 会尽可能地占据父容器在水平和垂直方向上的剩余空间
                                                                            #expand 参数决定组件是否会扩展以填充父容器中未被其他组件使用的额外空间。当 expand 设置为 True 时，self.timer_tree 会在父容器中有多余空间时进行扩展，进一步填充这些空间。
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)  #fill 参数用于指定组件在父容器中如何填充可用空间。tk.Y 表示组件将在垂直方向上填充可用空间，也就是 self.scrollbar 会在垂直方向上尽可能地占据父容器的高度，以适应 Treeview 表格可能的滚动需求

        '''在上半部分使用 left_frame 和 right_frame 来创建两个区域，是因为上半部分需要将按钮和备注输入区域进行分组和布局，使它们分别位于不同的区域，并且有不同的排列方式和间距要求。
            而对于下半部分的 tree 和 scrollbar，它们的布局相对简单，主要是将 tree 表格和与之关联的垂直滚动条组合在一起。具体原因如下：
            1.功能上的紧密关联：tree 和 scrollbar 是紧密相关的组件，scrollbar 是为了方便 tree 在内容过多时进行滚动查看而存在的，它们在功能上是一个整体，
            主要目的是展示和操作表格数据，不需要像上半部分那样进行复杂的分组和区域划分。
            2.布局需求简单：只需要将 tree 放置在左侧并让其在水平和垂直方向上填充空间，scrollbar 放置在右侧并在垂直方向上填充空间，这种布局方式较为直接和简单，不需要额外的框架来进行组织。
            不像上半部分的按钮和备注输入区域，有不同的对齐方式、间距要求以及可能的后续扩展需求，所以不需要创建额外的区域框架来管理它们的布局。'''

        # 绑定鼠标右键事件
        self.timer_tree.bind("<Button-3>", self.show_context_menu)

        # 创建上下文菜单
        self.context_menu = tk.Menu(root, tearoff=0)  #tearoff=0 表示这个菜单不能被从窗口中分离出来形成独立的窗口，让菜单的使用更符合常规的交互方式。即去掉了菜单顶部默认的虚线分割线
        self.context_menu.add_command(label="复制该行备注到输入框", command=lambda event=None: self.copy_note_to_entry(event))
                            # add_command此方法为菜单添加一个菜单项
                            # command=lambda event=None: self.copy_note_to_entry(event) ，当用户点击这个菜单项时，会调用 self.copy_note_to_entry方法，并将事件对象 event 作为参数传递给它。

        # 创建“是否自动开始计时”按钮
        self.auto_start_var = tk.BooleanVar()
        self.auto_start_var.set(False)
        self.auto_start_button = tk.Checkbutton(left_frame, text="是否自动开始计时", variable=self.auto_start_var)
        self.auto_start_button.pack(pady=5, anchor=tk.W)
        '''--------------------------------------------------------------------------------------------------------------------------------------------------------'''


        '''-------------------------------------------------------------------To Do List页签------------------------------------------------------------------------'''
        # 创建待办事项页签
        self.todo_frame = tk.Frame(self.notebook)
        self.notebook.add(self.todo_frame, text="待办列表")

        # 创建 Treeview 显示待办事项
        self.todo_tree = ttk.Treeview(self.todo_frame, columns=("序号", "待办事项"), show="headings")
        self.todo_tree.heading("序号", text="序号")
        self.todo_tree.heading("待办事项", text="待办事项")
        self.todo_tree.column("序号", width=40)
        self.todo_tree.column("待办事项", width=200)
        self.todo_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 插入序号 1 到 50
        for i in range(1, 51):
            self.todo_tree.insert("", "end", values=(i, ""))

        # 绑定 Treeview 的单元格编辑事件
        self.todo_tree.bind("<Button-1>", self.on_single_click)

        # 创建垂直滚动条
        self.todo_scrollbar = ttk.Scrollbar(self.todo_frame, orient=tk.VERTICAL, command=self.todo_tree.yview)
        self.todo_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        #将tree的垂直滚动操作关联到 scrollbar
        self.todo_tree.configure(yscrollcommand=self.todo_scrollbar.set)
        # 绑定鼠标滚轮事件
        self.todo_tree.bind("<MouseWheel>", self.on_mousewheel)

        # # 绑定右键菜单
        self.todo_tree.bind("<Button-3>", self.show_context_menu)

        # 绑定键盘事件
        self.todo_tree.bind("<Up>", self.on_up_key)
        self.todo_tree.bind("<Down>", self.on_down_key)
        self.todo_tree.bind("<Return>", self.on_return_key)

        self.update_note_combobox()

        # 调用导入 Excel 数据的方法
        self.load_latest_excel_data()
        '''---------------------------------------------------------------------------------------------------------------------------------------------------------'''

        '''----------------------------------------------------------------可视化页签----------------------------------------------------------------------------------'''
        # 可视化页签
        self.canvas_tab = tk.Frame(self.notebook)
        self.notebook.add(self.canvas_tab, text="可视化")
        
        # 初始化绘图（首次为空）
        self.draw_placeholder()
        '''---------------------------------------------------------------------------------------------------------------------------------------------------------'''


        '''----------------------------------------------------------------便签页签----------------------------------------------------------------------------------'''
        # 创建便签页签
        self.note_tab = tk.Frame(self.notebook)
        self.notebook.add(self.note_tab, text="便签")

        # 在便签页签中添加文本框
        self.note_text = tk.Text(self.note_tab, wrap=tk.WORD)
        self.note_text.pack(fill=tk.BOTH, expand=True)
        '''---------------------------------------------------------------------------------------------------------------------------------------------------------'''

        '''---------------------------------------------------------------使用说明页签-------------------------------------------------------------------------------'''
        # 创建使用说明页签
        self.contact_tab = tk.Frame(self.notebook)
        self.notebook.add(self.contact_tab, text="使用说明")

        # 在使用说明页签中添加不可编辑的文本标签
        contact_info = '''
        软件说明：\n
        1.点击完“暂停计时”后，需要点击“恢复计时”，否则到你点击“结束计时”的时候，会把你点击“暂停计时”之后的时间都计入“暂停时长”\n
        2.“暂停计时/恢复计时”可以多次点击，该部分的时间会累积计算\n
        3.“备注”是必填内容，否则无法点击“结束计时”，除非直接关闭该软件\n        
        4.当关闭该软件时，会自动导出一份Excel（主要是防止关闭前忘了按导出Excel按钮导致记录丢失）\n
        5.当点击了“开始计时”然后马上关闭该软件时。导出的Excel仅会记录之前已完成的行信息（如有），而最新一行的信息仅会记录“开始时间”而不会记录其他信息。\n
        6.点击“导出Excel”时仅会导出“计时器”页签下的内容，不会导出“便签”页签内的内容，请注意。\n
        7.在“计时器”页签下的表格中，点击右键可以复制之前的备注到输入框内。\n
        8.*V4*To Do List页签，用户可以在该页签下输入待办事项内容。且该页签的内容可以在“计时器”页签的备注中选取填入备注框\n
        9.*v5*To Do List页签，从双击激活单元格改为单击激活/ 可用上下方向键和回车键控制 / 导出Excel会连同待办事项一并导出到Excel单独的一个sheet上\n
        10.*v6*增加拖到到屏幕左右边缘自动隐藏的功能（多屏时副屏幕可能失效，可尝试拖动到不同屏幕的左右边缘测试）\n
        11.*v7*打开软件时会自动导入最近一份Excel（以Excel名判断）的待办事项到软件的待办列表内。 （注意：需要将代码和Excel放在同一个文件夹内）\n
        12.*v8*增加了设置页签，当点击结束计时后自动点击开始计时的选项（防止漏点） 
        13.*v9*删除设置页签，将自动开始计时选项移动到计时器页签
        14.*v9*增加了可视化页签，内涵饼图和柱状图（通过鼠标左键点击后切换），导出Excel时会一并将两个图表导出
        \n
        使用建议：\n
        1.*v9* 由于饼图会将具有同样备注的时间进行加总，而且当备注中间有“ - ”时会只取前半部分作为备注\n
        （例如“做报表 - 利润表”则只会取到“做报表”作为备注，并连同其他“做报表 - 资产表”的时间进行加总统计），因此可通过“ - ”来控制需要加总统计的项目\n
        \n
        如有任何建议请邮件联系作者。\n
        邮箱：tarry.zhou@outlook.com
        '''
        self.contact_label = tk.Label(self.contact_tab, text=contact_info, justify=tk.LEFT, padx=10, pady=10)
        self.contact_label.pack()
        '''---------------------------------------------------------------------------------------------------------------------------------------------------------'''


        '''---------------------------------------------------------------窗口吸附效果-------------------------------------------------------------------------------'''
        # 隐藏相关参数
        # 窗口边缘距离屏幕边缘小于此值时触发隐藏 （self.hide_threshold 不仅用于控制窗口何时隐藏，还用于控制鼠标靠近显示器边缘窗口弹出时边框距离屏幕边缘的距离。）
        self.hide_threshold = 2
        # 鼠标悬停检测范围（大于hide_threshold，主要是为了避免窗口在屏幕边缘频繁隐藏和显示（即 “闪烁” 问题））
        self.hover_threshold = 10
        # 窗口是否已隐藏的标志
        self.is_hidden = False
        # 窗口是否正在被拖动的标志
        self.is_dragging = False
        # 新增：记录上次隐藏的时间（时间戳）
        self.last_hide_time = 0
        # 冷却时间（毫秒），避免窗口高频显示和隐藏切换
        self.cool_down = 500

        # 绑定拖动事件
        # 当鼠标左键按下时，调用 start_drag 方法
        self.root.bind("<ButtonPress-1>", self.start_drag)
        # 当鼠标左键释放时，调用 stop_drag 方法
        self.root.bind("<ButtonRelease-1>", self.stop_drag)
        # 当鼠标左键按下并移动时，调用 on_drag 方法
        self.root.bind("<B1-Motion>", self.on_drag)

        # 每隔 50 毫秒调用一次 check_position 方法
        self.root.after(50, self.check_position)
        '''---------------------------------------------------------------------------------------------------------------------------------------------------------'''

        # 绑定窗口关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)


    '''----------------------------------------------------------计时器页签部分--------------------------------------------------------------------------------'''
    def start_timer(self):
        if self.start_time is None:        #若 self.start_time 为 None，则意味着计时还未开始，此时就会执行 if 语句块里的代码
            self.start_time = time.time()  #time.time() 返回的是从 1970 年 1 月 1 日午夜（UTC）开始到现在所经过的秒数，是一个浮点数
            start_time_str = time.strftime("%H:%M:%S", time.localtime(self.start_time)) #time.localtime 把self.start_time这个时间戳转换成本地时间的struct_time 对象。即是返回一个包含年月日等多个信息的元组
                                                                                        #time.strftime 它的作用是根据指定的格式字符串把 struct_time 对象格式化为字符串
            self.timer_tree.insert("", 0, values=(self.record_index, start_time_str, "", "", "", "")) #第一个参数""，代表父项的标识符。当传入空字符串 "" 时，意味着要插入的行是顶级行，也就是不在任何子树结构下的行
                                                                                                #第二个参数 0，指定了新插入行的位置。0 表示将新行插入到表格的第一行，也就是最顶部的位置。如果传入 "end"，则表示将新行插入到表格的最后一行。
                                                                                                #第三个参数values，用于指定新插入行各列的值。values 是一个元组，元组中的每个元素对应表格中的一列。
            self.record_index += 1  #每次运行后自增1
            self.start_button.config(state=tk.DISABLED)  #开始计时按钮变为禁用状态。即config 方法允许你在程序运行时改变组件的各种属性
            self.pause_button.config(state=tk.NORMAL)    #暂停按钮变为启用状态
            self.end_button.config(state=tk.NORMAL)      #结束按钮变为启用状态
            self.export_button.config(state=tk.DISABLED) #导出Excel按钮变为禁用状态

    def pause_timer(self):
        if not self.is_paused:  #如果self.is_paused为True（运行状态），那么not self.is_paused的结果就是False；
                                #如果self.is_paused为False（初始状态），那么not self.is_paused的结果就是True，那么就运行if内部代码
            self.pause_start_time = time.time()
            self.pause_button.config(text="继续计时", bg="green", fg="white")  #config 是 tkinter 组件提供的一个方法，用于在程序运行时修改组件的属性,通过传入不同的参数值可以改变组件的外观和行为。
            self.is_paused = True
        else:
            pause_end_time = time.time()
            self.pause_total_time += pause_end_time - self.pause_start_time
            self.pause_button.config(text="暂停计时", bg="red", fg="white")
            self.is_paused = False

    def end_timer(self):
        note = self.note_entry.get() #获取entry输入框内的内容
        if not note.strip():         # 检查备注是否为空
            messagebox.showwarning("提示", "请先填写备注再点结束计时！")
            return                   # 若备注为空，直接返回，不执行后续计时结束操作

        if self.start_time is not None:
            end_time = time.time()
            if not self.is_paused:     #如果self.is_paused为False（初始状态），那么not self.is_paused的结果就是True，那么就运行if内部代码
                total_duration = end_time - self.start_time - self.pause_total_time
            else:                      #如果self.is_paused为True（运行状态），那么not self.is_paused的结果就是False，那么就运行else内部代码
                pause_end_time = end_time
                self.pause_total_time += pause_end_time - self.pause_start_time  #这句是防止点了暂停计时后未点恢复计时就直接点结束计时，避免遗漏计算暂停时长
                total_duration = end_time - self.start_time - self.pause_total_time

            end_time_str = time.strftime("%H:%M:%S", time.localtime(end_time))  #转换结束时间格式
            pause_total_str = self.format_time(self.pause_total_time)           #将暂停时长（即结束-开始的浮点数）转换格式，若大于60秒就转成分钟
            total_duration_str = self.format_time(total_duration)               #将用工时长（即结束-开始的浮点数）转换格式，若大于60秒就转成分钟      

            # 更新表格第一行数据
            item = self.timer_tree.get_children()[0]                                  #get_children()作用是返回一个包含当前 Treeview 中所有子项（即表格中的行）标识的元组。索引 [0] 表示从这个元组中取出第一个元素。在表格的上下文中，这意味着获取表格中的第一行的标识
            self.timer_tree.item(item, values=(self.record_index - 1,                 #因为在点完“开始计时”之后，record_index就自增1了，所以要用-1来找回已经填写了开始时间的那行。
                                        time.strftime("%H:%M:%S", time.localtime(self.start_time)),
                                        end_time_str, pause_total_str, total_duration_str, note))
            self.reset_timer()
            self.export_button.config(state=tk.NORMAL)
            self.check_and_update_todo_list()

            # 检查是否自动开始计时
            if self.auto_start_var.get():
                self.start_timer()
            
            self.get_latest_analysis_data()

    def reset_timer(self):
        self.start_time = None
        self.pause_start_time = None
        self.pause_total_time = 0
        self.is_paused = False
        self.start_button.config(state=tk.NORMAL)
        self.pause_button.config(text="暂停计时", bg="red", fg="white")
        self.pause_button.config(state=tk.DISABLED)
        self.end_button.config(state=tk.DISABLED)
        self.note_entry.delete(0, tk.END)

    # def export_to_excel(self):
        # # 检查表格是否有数据
        # timer_items = self.timer_tree.get_children()
        # if not timer_items:
        #     messagebox.showwarning("警告", "计时器页签表格没有数据，无法导出。")
        #     return

        # # 获取计时器页签数据
        # timer_data = []
        # for item in timer_items:
        #     values = self.timer_tree.item(item)["values"]
        #     timer_data.append(values)

        # # 获取待办列表页签数据
        # todo_items = self.todo_tree.get_children()
        # todo_data = []
        # for item in todo_items:
        #     values = self.todo_tree.item(item)["values"]
        #     todo_data.append(values)

        # # 设置Excel名称
        # date_str = time.strftime("%Y-%m-%d", time.localtime())
        # start_time_str = self.timer_tree.item(self.timer_tree.get_children()[-1], 'values')[1].replace(":", "")
        # end_time_str = self.timer_tree.item(self.timer_tree.get_children()[0], 'values')[2].replace(":", "")
        # file_name = f"计时器：{date_str} {start_time_str}-{end_time_str}.xlsx"

        # # 创建Excel文件
        # with pd.ExcelWriter(file_name) as writer:
        #     # 导出计时器页签数据到第一个sheet
        #     df_timer = pd.DataFrame(timer_data, columns=["序号", "开始时间", "结束时间", "暂停时长", "用工时长", "备注"])
        #     df_timer.to_excel(writer, sheet_name='计时器页签', index=False)

        #     # 导出待办列表页签数据到第二个sheet
        #     df_todo = pd.DataFrame(todo_data, columns=["序号", "待办事项"])
        #     df_todo.to_excel(writer, sheet_name='待办列表页签', index=False)

        # messagebox.showinfo("提示", f"数据已成功导出到{file_name}")

    def export_to_excel(self):
        # 检查表格是否有数据
        timer_items = self.timer_tree.get_children()
        if not timer_items:
            messagebox.showwarning("警告", "计时器页签表格没有数据，无法导出。")
            return

        # 获取计时器页签数据
        timer_data = []
        for item in timer_items:
            values = self.timer_tree.item(item)["values"]
            timer_data.append(values)

        # 获取待办列表页签数据
        todo_items = self.todo_tree.get_children()
        todo_data = []
        for item in todo_items:
            values = self.todo_tree.item(item)["values"]
            todo_data.append(values)

        # 设置 Excel 名称
        date_str = time.strftime("%Y-%m-%d", time.localtime())
        start_time_str = self.timer_tree.item(self.timer_tree.get_children()[-1], 'values')[1].replace(":", "")
        end_time_str = self.timer_tree.item(self.timer_tree.get_children()[0], 'values')[2].replace(":", "")
        file_name = f"计时器：{date_str} {start_time_str}-{end_time_str}.xlsx"

        # 创建 Excel 文件
        wb = Workbook()

        # 导出计时器页签数据到第一个 sheet
        ws1 = wb.active
        ws1.title = '计时器页签'
        df_timer = pd.DataFrame(timer_data, columns=["序号", "开始时间", "结束时间", "暂停时长", "用工时长", "备注"])
        for r_idx, row in enumerate(df_timer.values.tolist(), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws1.cell(row=r_idx, column=c_idx, value=value)
        # 写入表头
        for col_idx, col_name in enumerate(df_timer.columns, start=1):
            ws1.cell(row=1, column=col_idx, value=col_name)

        # 导出待办列表页签数据到第二个 sheet
        ws2 = wb.create_sheet('待办列表页签')
        df_todo = pd.DataFrame(todo_data, columns=["序号", "待办事项"])
        for r_idx, row in enumerate(df_todo.values.tolist(), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws2.cell(row=r_idx, column=c_idx, value=value)
        # 写入表头
        for col_idx, col_name in enumerate(df_todo.columns, start=1):
            ws2.cell(row=1, column=col_idx, value=col_name)

        # 创建第三个 sheet 用于存放图表
        ws3 = wb.create_sheet('图表')

        # 将饼图插入到第三个 sheet
        if self.pie_img:
            # 新增：重新创建一个新的 BytesIO 对象
            new_pie_img = BytesIO(self.pie_img.getvalue())
            img_pie_obj = Image(new_pie_img)
            ws3.add_image(img_pie_obj, 'A1')

        # 将柱状图插入到第三个 sheet
        if self.bar_img:
            # 新增：重新创建一个新的 BytesIO 对象
            new_bar_img = BytesIO(self.bar_img.getvalue())
            img_bar_obj = Image(new_bar_img)
            ws3.add_image(img_bar_obj, 'A35')  # 调整柱状图的插入位置

        # 保存 Excel 文件
        wb.save(file_name)
        messagebox.showinfo("提示", f"数据已成功导出到{file_name}")


    def format_time(self, seconds):
        if seconds >= 60:
            minutes = seconds / 60
            return f"{minutes:.2f} 分钟"
        return f"{seconds:.2f} 秒"

    def on_mousewheel(self, event):
        self.timer_tree.yview_scroll(int(-1 * (event.delta / 120)), "units")  #yview_scroll 是 Treeview 组件的一个方法，用于在垂直方向上滚动表格的视图。它接受两个参数
                                                                        #第一个参数是滚动的单位数量：
                                                                        #鼠标滚轮每滚动一次，delta 的值通常为 120 或 -120，正数表示向上滚动，
                                                                        #将 delta 的值除以 120，得到鼠标滚轮滚动的单位数。
                                                                        #乘以 -1 是为了调整滚动方向，使得滚动方向与鼠标滚轮的实际滚动方向一致
                                                                        #第二个参数是滚动的单位类型：
                                                                        #这里是 "units"，表示按单位滚动，一个单位通常对应表格中的一行

    def on_close(self):
        # 在关闭窗口时导出 Excel
        self.export_to_excel()
        self.root.destroy()

    def show_context_menu(self, event):
        # 显示上下文菜单
        self.context_menu.post(event.x_root, event.y_root) #event.x_root 是鼠标点击位置相对于屏幕左上角的x坐标。
        self.context_menu.event = event  # 将当前的事件对象event保存为 self.context_menu的一个属性event，给后续的copy_note_to_entry方法中使用

    # 计时器页签下的“复制备注到输入框”功能
    def copy_note_to_entry(self, event=None):
        if event is None:    #检查传入的事件对象 event 是否为 None。如果是，则从 self.context_menu.event 中获取之前保存的事件对象
            event = self.context_menu.event
        # 获取鼠标所在的行
        item = self.timer_tree.identify_row(event.y)  #identify_row 是 Treeview 组件的一个方法，用于根据给定的垂直坐标（相对于 Treeview 组件本身）来确定鼠标所在的表格行。
                                                #event.y 是鼠标点击位置相对于 Treeview 组件左上角的垂直坐标。通过这个坐标，identify_row 方法会返回鼠标点击位置所在行的标识符（通常是一个字符串），
                                                #将其赋值给变量 item，后续就可以使用这个标识符来获取该行的具体数据。
        if item:
            # 获取该行的备注内容
            note = self.timer_tree.item(item, "values")[-1] # self.timer_tree.item(item, "values") 用于获取指定行（由 item 标识）的所有列数据，返回一个元组。由于备注在最后一列，因此用-1获取
            # 将内容输入到备注输入框内
            self.todo_var.set(note)

    '''------------------------------------------------------------------------------------------------------------------------------------------------------'''


    '''----------------------------------------------------------待办事项页签部分--------------------------------------------------------------------------------'''
    # 定义鼠标单击待办列表单元格后的处理
    def on_single_click(self, event):
        # 确保 Treeview 获得焦点
        self.todo_tree.focus_set()            #在图形用户界面（GUI）里，焦点指的是当前接收用户输入的组件。当一个组件拥有焦点时，用户的键盘输入（像按键、输入文本等）会直接作用于该组件。  
        # 单击 Treeview 单元格时进入编辑状态
        region = self.todo_tree.identify_region(event.x, event.y)
        if region == "cell":                                     #判断点击区域是否为单元格
            item = self.todo_tree.identify_row(event.y)          #根据鼠标点击的纵坐标 event.y 确定点击的是 Treeview 中的哪一行，返回该行的标识符 item
            col = self.todo_tree.identify_column(event.x)        #根据鼠标点击的横坐标 event.x 确定点击的是 Treeview 中的哪一列，返回列的标识符 col
            if col == "#2":                                      #只允许编辑第二列 
                self.create_entry(item, col)                     

    # 处理在Treeview 本身的方向键事件（而不是输入框的上下方向键），即当焦点在 Treeview 上（未进入单元格编辑状态，仅选中某一行）时，按下上下方向键会触发这两个方法，用于切换选中的行（不涉及输入框的编辑操作）。
    def on_up_key(self, event):
        # 处理 Treeview 向上方向键：切换选中行（未进入编辑状态时）
        adjacent_item = self.get_adjacent_item(-1)  # -1 表示向上
        if adjacent_item:
            self.todo_tree.selection_set(adjacent_item)

    # 处理在Treeview 本身的方向键事件（而不是输入框的上下方向键），即当焦点在 Treeview 上（未进入单元格编辑状态，仅选中某一行）时，按下上下方向键会触发这两个方法，用于切换选中的行（不涉及输入框的编辑操作）。
    def on_down_key(self, event):
        # 处理 Treeview 向下方向键：切换选中行（未进入编辑状态时）
        adjacent_item = self.get_adjacent_item(1)  # 1 表示向下
        if adjacent_item:
            self.todo_tree.selection_set(adjacent_item)

    def on_return_key(self, event):
        selected_item = self.todo_tree.selection()
        if selected_item:
            item = selected_item[0]
            col = "#2"
            for widget in self.todo_tree.winfo_children():
                if isinstance(widget, ttk.Entry):
                    self.save_value(event, item, col, widget)
                    self.handle_entry_up_down(event, item, 1, None)

    # 根据传入的方向（向上或向下），在 Treeview 中找到当前选中项的相邻项的标识符
    def get_adjacent_item(self, direction):
        selected_item = self.todo_tree.selection()      # 获取当前 Treeview 中被选中的项，返回值是一个包含选中项标识符的元组
        if not selected_item:                           # 检查是否有项被选中，如果没有选中项，则直接返回 None
            return None                                 
        index = self.todo_tree.index(selected_item[0])  # 由于 selected_item 是元组，取其第一个元素作为当前选中项的标识符
        children = self.todo_tree.get_children()        # 获取 Treeview 中所有项的标识符列表
        new_index = index + direction                   # 根据传入的方向参数计算相邻项的索引，如果 direction 为 -1 则计算上一项的索引，如果为 1 则计算下一项的索引
        if 0 <= new_index < len(children):              # 检查新计算出的索引是否在有效范围内，有效范围是索引大于等于 0 且小于 Treeview 中项的总数
            return children[new_index]                  # 如果新索引有效，从所有项的标识符列表中获取对应的相邻项标识符并返回
        return None                                     # # 如果新索引越界，即超出了 Treeview 中项的范围，返回 None

    # 处理输入框上下方向键（而不是Treeview的上下方向键）事件并切换编辑单元格
    def handle_entry_up_down(self, event, current_item, direction, entry):    #形参包括：self代表类的实例对象，通过它可以访问类的属性和其他方法。 
                                                                                        #event触发此方法的事件相关信息，比如按键事件的具体按键等
                                                                                        #current_item表示当前正在编辑的 Treeview 行的标识符
                                                                                        #direction一个整数，-1 表示向上移动，1 表示向下移动
                                                                                        #entry当前正在编辑的 ttk.Entry 输入框对象，用于获取和保存用户输入的内容
        item = self.todo_tree.selection()[0]          #获取当前被选中的 Treeview 行的标识符。因为 selection() 方法返回的是一个包含所有选中行标识符的元组，这里假设只有一行被选中，所以取索引为 0 的元素
        col = "#2"                                    #只允许编辑第二列 
        if entry:                                     #检查 entry 是否存在（不为 None）。如果存在，则调用 self.save_value 方法保存当前输入框中的内容。save_value 方法通常会将输入框中的内容更新到 Treeview 对应的单元格中，并销毁该输入框。
            self.save_value(event, item, col, entry)  # 保存当前输入框内容
        new_item = self.get_adjacent_item(direction)  # direction 直接传递（-1 或 1）
        if new_item:                                  # 仅当相邻行存在时（非越界）
            self.todo_tree.selection_set(new_item)    # 切换选中行
            self.create_entry(new_item, col)          # 创建新输入框

    # 创建并初始化用于单元格编辑的输入框
    def create_entry(self, item, col):                        
        x, y, width, height = self.todo_tree.bbox(item, col)    # 获取新单元格位置和大小：获取指定行和列的单元格在屏幕上的位置和大小，返回值分别是单元格左上角的横坐标 x、纵坐标 y 以及宽度 width 和高度 height
        value = self.todo_tree.set(item, col)                   # 获取新单元格当前内容：获取指定行和列的单元格的当前内容
        entry = ttk.Entry(self.todo_tree)                       # 创建新输入框：创建一个 Entry 输入框组件，用于用户输入新的内容
        entry.place(x=x, y=y, width=width, height=height)       # 放置新输入框：将输入框放置在被点击的单元格的位置，使其覆盖原单元格。
        entry.insert(0, value)                                  # 插入新内容：将原单元格的内容插入到输入框中，方便用户查看和修改。
        entry.focus()                                           # 设置焦点：将焦点设置到输入框上，使用户可以直接开始输入。
        # 绑定上下方向键和回车键事件到 Entry
        entry.bind("<Up>", lambda e: self.handle_entry_up_down(e, item, -1, entry))
        entry.bind("<Down>", lambda e: self.handle_entry_up_down(e, item, 1, entry))
        entry.bind("<Return>", lambda e: self.handle_entry_return(e, item, col, entry))
        entry.bind("<FocusOut>", lambda e: self.save_value(e, item, col, entry))

    def handle_entry_return(self, event, item, col, entry):
        # 保存当前输入框内容
        self.save_value(event, item, col, entry)
        # 模拟按下下方向键
        self.handle_entry_up_down(event, item, 1, None)

    def save_value(self, event, item, col, entry):
        if entry and entry.winfo_exists():           # 检查 Entry 组件是否存在
            new_value = entry.get()
            self.todo_tree.set(item, col, new_value) #将新内容保存到 Treeview 中指定行和列的单元格中。
            entry.destroy()                          #销毁输入框，恢复 Treeview 的显示

    def update_note_combobox(self):
        # 获取所有待办事项
        items = self.todo_tree.get_children()
        todo_list = []
        for item in items:
            todo = self.todo_tree.item(item)["values"][1]
            if todo:
                todo_list.append(todo)
        self.note_entry['values'] = todo_list

    def update_combobox_on_click(self, event):
        self.update_note_combobox()

    # 检查计时器页签的备注列内容是否在待办列表页签的待办事项列中，如果有重复的内容，则删除待办列表页签中的重复内容，并将其后的内容上移补位
    def check_and_update_todo_list(self):
        # 获取计时器页签的备注列内容
        timer_notes = [self.timer_tree.item(item)["values"][-1] for item in self.timer_tree.get_children()]

        # 获取待办列表页签的待办事项列内容
        todo_items = [self.todo_tree.item(item)["values"][-1] for item in self.todo_tree.get_children()]

        # 移除 todo_items 中与 timer_notes 重复的元素，同时保持顺序
        temperary_list = []
        for item in todo_items:
            if item not in timer_notes:
                temperary_list.append(item)

        # 计算 non_empty_list 的长度，并用空值在其后补足至 50 个元素
        if len(temperary_list) < 50:
            temperary_list += [""] * (50 - len(temperary_list))

        # 清空 Treeview 的内容
        for item in self.todo_tree.get_children():
            self.todo_tree.delete(item)

        # 将 non_empty_list 的内容更新到待办列表页签的表格中，序号从 1 开始
        for i, item in enumerate(temperary_list, start=1):
            self.todo_tree.insert("", "end", values=(i, item))

    # 读取exe同目录下最新的 Excel 文件，并将“待办事项”列的内容填入 Treeview 表格
    def load_latest_excel_data(self):
        # 获取程序运行时的 **当前工作目录**（最可靠，无论是否打包）
        current_dir = os.getcwd()
        # 存储以“计时器”开头的 Excel 文件及其日期时间
        excel_files = []
        # 遍历当前目录下的所有文件
        for filename in os.listdir(current_dir):                  #os.listdir(current_dir)：返回当前目录下的所有文件和文件夹的名称列表
            if filename.startswith("计时器") and filename.endswith(".xlsx"):
                try:
                    # 提取文件名中的日期时间部分
                    date_str = filename.split("：")[1].split(".xlsx")[0]  #第一次拆分是获取冒号后面的字节，第二次拆分是获取.xlsx前面的字节
                    # 将日期时间字符串转换为 datetime 对象
                    date = datetime.strptime(date_str, "%Y-%m-%d %H%M%S-%f")
                    excel_files.append((date, os.path.join(current_dir, filename)))  #将日期，Excel文件的绝对路径打包成元组存入excel_files列表
                except (IndexError, ValueError):
                    continue
        # 如果没有找到符合条件的文件，直接返回
        if not excel_files:
            return
        # 按日期时间降序排序
        excel_files.sort(reverse=True)
        # 获取最新的 Excel 文件路径
        latest_file = excel_files[0][1]  #获取排序后列表中第一个元素（元组），的第二个元素，即最新文件的完整路径。
        try:
            # 读取 Excel 文件中的“待办列表页签”工作表
            df = pd.read_excel(latest_file, sheet_name="待办列表页签")
            # 获取第二列“待办事项”从第 2 行开始的内容
            todo_items = df.iloc[0:, 1].dropna().tolist()  # 去除空值,并转换成列表  #因为第一行已经被pd认定为是表头，所以第一行相当于Excel本身的第二行，所以用下标0
            # 将待办事项填入 Treeview 表格的第二列
            for i, item in enumerate(todo_items):
                if i < 50:
                    self.todo_tree.set(self.todo_tree.get_children()[i], "待办事项", item) #使用 Treeview 的 set 方法将待办事项填入表格的第二列（“待办事项” 列）
        except Exception as e:
            print(f"读取 Excel 文件时出错: {e}")
    '''-----------------------------------------------------------------------------------------------------------------------------------------------------'''


    '''----------------------------------------------------------边缘吸附效果--------------------------------------------------------------------------------'''
    def start_drag(self, event):
        # 鼠标左键按下，设置拖动状态为 True
        self.root.is_dragging = True

    def stop_drag(self, event):
        # 鼠标左键释放，设置拖动状态为 False
        self.root.is_dragging = False

    def on_drag(self, event):
        # 拖动时无需操作，由系统自动处理窗口位置
        pass

    def get_window_and_mouse_info(self):
        # 获取窗口的当前 x 和 y 坐标
        x, y = self.root.winfo_x(), self.root.winfo_y()                   
        # 获取窗口的宽度和高度
        width, height = self.root.winfo_width(), self.root.winfo_height() 
        # 获取鼠标的当前 x 和 y 坐标
        mouse_x, mouse_y = self.root.winfo_pointerx(), self.root.winfo_pointery() 
        # 返回窗口的 x坐标，y坐标，宽度，高度   以及鼠标的 x坐标，y坐标
        return x, y, width, height, mouse_x, mouse_y

    def is_mouse_in_window(self):
        x, y, width, height, mouse_x, mouse_y = self.get_window_and_mouse_info()
        return x <= mouse_x <= (x + width) and y <= mouse_y <= (y + height)

    def check_position(self):
        #因为get_window_and_mouse_info返回的还会有鼠标的y坐标，但用不上，所以最后一个用 _ 来接收返回值
        x, y, width, height, mouse_x, _ = self.get_window_and_mouse_info()  
        # 获取屏幕的宽度和高度
        screen_width, screen_height = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        # 获取当前时间（毫秒）
        current_time = time.time() * 1000

        # ------------------- 窗口处于隐藏隐藏状态，需要弹出时的处理 -------------------
        if self.is_hidden:
            # 冷却时间内不允许显示（避免高频切换）  此条件判断当前时间距离窗口上次隐藏的时间是否小于冷却时间。如果小于，说明还处于冷却期，窗口不允许重新显示。
            if current_time - self.last_hide_time < self.cool_down:
                # 继续等待 50 毫秒后再次检查位置
                self.root.after(50, self.check_position)
                # 跳过本次检测
                return

            # 判断鼠标是否悬停在窗口左侧检测范围内，如果是将 is_hover_left 设为 True，否则设为 False
            is_hover_left = mouse_x < self.hover_threshold
            # 判断鼠标是否悬停在窗口右侧检测范围内，如果是将 is_hover_left 设为 True，否则设为 False。因为hover_threshold是距离值，而鼠标需要停留在右边缘的实际位置是右屏幕边缘-距离值
            is_hover_right = mouse_x > (screen_width - self.hover_threshold)

            if is_hover_left or is_hover_right:  # 当鼠标在左右悬停区域内时
                # 计算窗口弹出时，窗口左边框所处的屏幕X轴（横向）位置。
                # 如果鼠标是在左悬停区则窗口弹出时，就显示在距离值（hover_threshold -1）的位置。此时距离值相当于屏幕横向x轴位置，因此无需再计算
                # 如果鼠标是在右悬停区则窗口弹出时，就显示在（屏幕左右总宽度 - 窗口宽度 - 距离值-1）的位置
                # 因为当窗口边缘小于距离值的位置时，才会被自动吸附，因此用 hover_threshold -1
                target_x = (self.hide_threshold -1) if is_hover_left else screen_width - width - (self.hide_threshold-1)
                # 设置窗口的大小和位置
                self.root.geometry(f"{width}x{height}+{target_x}+{y}")
                # 设置窗口为可见状态
                self.is_hidden = False

        # ------------------- 窗口处于可见状态，需要隐藏时的处理 -------------------
        else:
            # 判断窗口是否在屏幕边缘，x是窗口当前的水平坐标（窗口左上角的 x 坐标）
            is_at_edge = (x < self.hide_threshold) or (x > screen_width - width - self.hide_threshold)
            # 判断鼠标是否在窗口内
            is_mouse_outside = not self.is_mouse_in_window()
            # 判断鼠标是否在悬停范围外
            is_mouse_out_of_hover = not (mouse_x < self.hover_threshold or mouse_x > (screen_width - self.hover_threshold))

            # 只有当窗口在屏幕边缘、鼠标不在窗口内且不在悬停范围内，同时窗口没有被拖动时，才隐藏窗口
            if is_at_edge and is_mouse_outside and is_mouse_out_of_hover and not self.is_dragging:
                # 隐藏窗口并记录隐藏时间
                # 根据窗口位置确定隐藏后的 x 坐标
                self.root.geometry(f"{width}x{height}+{-width if x < self.hide_threshold else screen_width}+{y}") #窗口宽度*高度 + {窗口负宽度 if x小于左距离值 else 将窗口放在屏幕右边界外} + 窗口高度
                # 设置窗口为隐藏状态
                self.is_hidden = True
                # 记录隐藏时间
                self.last_hide_time = current_time

        # 继续等待 50 毫秒后再次检查位置
        self.root.after(50, self.check_position)
    '''---------------------------------------------------------------------------------------------------------------------------------------------------------'''

    '''----------------------------------------------------------可视化页签--------------------------------------------------------------------------------'''    
    def draw_placeholder(self):
        """带数据提示的占位图"""
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.text(0.5, 0.5, 
                "请先在【计时】页签记录数据\n\n点击此处开始生成图表", 
                ha="center", va="center", 
                fontsize=12, 
                bbox=dict(facecolor='white', alpha=0.8))  # 添加背景框增强可读性
        ax.axis("off")
        
        self.canvas = FigureCanvasTkAgg(fig, master=self.canvas_tab)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack()
        
        # 绑定点击事件（仅在数据存在时切换）
        self.canvas.mpl_connect('button_press_event', self.on_placeholder_click)

    def on_placeholder_click(self, event):
        """占位图点击处理（区分数据是否存在）"""
        analysis_list = self.get_latest_analysis_data()
        if len(analysis_list) == 0:
            messagebox.showinfo("提示", "请先记录至少一条计时数据")
        else:
            self.current_chart = "pie"  # 首次默认显示饼图
            self.draw_maps(analysis_list)

    def draw_maps(self, analysis_list):
        """主绘图方法（增强初始状态处理）"""
        if not analysis_list:
            self.draw_placeholder()
            return  # 数据为空时不执行后续绘图
        
        df = self.process_data(analysis_list)
        
        # 首次绘制时根据current_chart状态选择图表（默认饼图）
        if self.current_chart not in ["pie", "bar"]:
            self.current_chart = "pie"  # 初始化图表类型
        
        # 销毁旧画布（确保每次绘制都是全新的）
        if self.canvas:
            self.canvas.get_tk_widget().destroy()
        
        # 根据当前类型绘制
        if self.current_chart == "pie":
            self.draw_pie_chart(df)
        else:
            self.draw_bar_chart(df)
        
        # 绑定事件（点击切换图表类型）
        self.canvas.mpl_connect('button_press_event', self.toggle_chart)

    def process_data(self, analysis_list):
        """统一数据处理方法（关键步骤）"""
        df = pd.DataFrame(analysis_list, columns=["序号", "开始时间", "结束时间", "暂停时长", "用工时长", "备注"])
        df['备注切片'] = df['备注'].str.split(' - ').str[0]

        # 用工时长处理（转换为数值型分钟）
        def convert_to_minutes(value):
            if isinstance(value, str):
                if value.endswith(' 秒'):
                    try:
                        return float(value.replace(' 秒', '')) / 60
                    except ValueError:
                        return None
                elif value.endswith(' 分钟'):
                    try:
                        return float(value.replace(' 分钟', ''))
                    except ValueError:
                        return None
            return value

        df['用工时长'] = df['用工时长'].apply(convert_to_minutes)
        df['用工时长'] = pd.to_numeric(df['用工时长'], errors='coerce')

        # 修正此处，正确使用 apply 方法
        df['用工小时'] = df['用工时长'].apply(lambda x: f'{round(x / 60, 2)}小时' if pd.notna(x) else None)
        df = df.sort_values('序号', ascending=True)
        return df

    def draw_pie_chart(self, df):
        """绘制饼图（修正分组数据问题）"""
        # 分组统计（仅数值列，不包含字符串列）
        pie_df = df.groupby('备注切片')['用工时长'].sum().reset_index()  # 新增reset_index()
        pie_df = pie_df.sort_values('用工时长', ascending=False)

        if pie_df.empty:
            self.draw_placeholder()
            return

        # 重新计算用工小时（基于分组后的用工时长）
        pie_df['用工小时'] = pie_df['用工时长'].apply(lambda x: f'{round(x / 60, 2)}小时')

        fig, ax = plt.subplots(figsize=(10, 6))  # 增加图形宽度
        explode = [0.05 if i == 0 else 0 for i in range(len(pie_df))]

        ax.pie(pie_df['用工时长'],
            labels=pie_df['备注切片'],  # 标签使用备注切片
            autopct='%.2f%%',
            textprops={'fontsize': 12},
            explode=explode)
        ax.set_title('任务时长占比', fontsize=14)
        ax.legend(pie_df['用工小时'], bbox_to_anchor=(1.2, 0.5), loc='center left', title='时长')  # 调整图例位置

        self.canvas = FigureCanvasTkAgg(fig, master=self.canvas_tab)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack()

        # 保存饼图为内存中的图像
        self.pie_img = BytesIO()
        fig.savefig(self.pie_img, format='png')
        self.pie_img.seek(0)
        plt.close(fig)

    # def draw_bar_chart(self, df):
    #     """绘制柱状图"""
    #     fig, ax = plt.subplots(figsize=(10, 6))  # 增加图形宽度
    #     x = df['开始时间']
    #     y = df['用工时长']

    #     ax.bar(x, y, color='darkorange', edgecolor='white')
    #     ax.set_title('各任务开始时间及时长', fontsize=14)
    #     ax.set_xlabel('开始时间', fontsize=12)
    #     ax.set_ylabel('分钟', fontsize=12)

    #     # 添加数据标签
    #     for i, (a, b) in enumerate(zip(x, y)):
    #         ax.text(a, b, f'{b:.2f}', ha='center', va='bottom', fontsize=10)

    #     plt.xticks(rotation=45, ha='right', fontsize=10)  # 旋转 x 轴标签
    #     plt.subplots_adjust(bottom=0.2)  # 调整子图布局，增加底部空间

    #     self.canvas = FigureCanvasTkAgg(fig, master=self.canvas_tab)
    #     self.canvas.draw()
    #     self.canvas.get_tk_widget().pack()

    def draw_bar_chart(self, df):
        """绘制柱状图（包含备注切片且按序号排序，xlabel 显示开始时间，柱体文字黑色）"""
        # 将序号转换为数字类型
        df['序号'] = pd.to_numeric(df['序号'], errors='coerce')
        # 按序号排序
        df = df.sort_values('序号', ascending=True)

        fig, ax = plt.subplots(figsize=(12, 6))
        x = df['序号']  # 使用序号作为绘图的 x 坐标
        y = df['用工时长']
        labels = df['备注切片']
        start_times = df['开始时间']  # 获取开始时间

        # 绘制柱状图
        bars = ax.bar(x, y, color='darkorange', edgecolor='white', alpha=0.8)

        ax.set_title('各任务开始时间及时长', fontsize=14)
        ax.set_xlabel('开始时间', fontsize=12)
        ax.set_ylabel('分钟', fontsize=12)

        # 添加备注切片到柱体内部，字体颜色设为黑色
        for bar, label in zip(bars, labels):
            height = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                height * 0.9,
                label,
                ha='center',
                va='center',
                fontsize=9,
                color='blue',  # 设置字体颜色为黑色
                rotation = 90
            )

        # 添加数据标签（用工时长）到柱体上方
        for a, b in zip(x, y):
            ax.text(a, b, f'{b:.2f}', ha='center', va='bottom', fontsize=10, fontweight='bold')

        # 设置 x 轴刻度为序号，标签显示为开始时间
        ax.set_xticks(x)
        ax.set_xticklabels(start_times)
        plt.xticks(rotation=45, ha='right', fontsize=10)
        plt.subplots_adjust(bottom=0.2, left=0.1, right=0.9, top=0.9)

        self.canvas = FigureCanvasTkAgg(fig, master=self.canvas_tab)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack()

        # 保存柱状图为内存中的图像
        self.bar_img = BytesIO()
        fig.savefig(self.bar_img, format='png')
        self.bar_img.seek(0)
        plt.close(fig)


    def toggle_chart(self, event):
        """鼠标左键点击切换图表类型"""
        if event.button == 1:  # 左键点击
            self.current_chart = "bar" if self.current_chart == "pie" else "pie"
            # 重新绘制图表（需获取最新数据，这里假设analysis_list可通过方法获取）
            # 注意：这里需要从treeview重新获取最新数据，不能直接使用历史数据
            analysis_list = self.get_latest_analysis_data()
            self.draw_maps(analysis_list)

    def get_latest_analysis_data(self):
        """从Treeview获取数据（修复可能的列索引错误）"""
        analysis_list = []
        item_ids = self.timer_tree.get_children()
        for item_id in item_ids:
            # 确保values包含正确的列（根据Treeview定义顺序）
            values = self.timer_tree.item(item_id, 'values')
            # 检查是否包含所有必要列（序号、开始时间、用工时长、备注等）
            if all(values):  # 检查是否有空值
                analysis_list.append(values)
        return analysis_list

    '''---------------------------------------------------------------------------------------------------------------------------------------------------------'''


if __name__ == "__main__":
    root = tk.Tk()
    root.geometry('800x200+600+300') #长*高+软件初始显示的位置
    app = TimerApp(root)
    root.mainloop()